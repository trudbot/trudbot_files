import openpyxl
from openpyxl import utils
from openpyxl import styles

'''打开一个表格'''
excel = openpyxl.load_workbook('第三次作业批改.xlsx', data_only=True)  # 打开表格，并创建一个workbook对象
# data_only用于读取cell中的值，当单元格中的值是一个公式的时候，会返回计算到的结果。
print(type(excel))

sheet_active = excel.active  # 得到workbook的活动表， 即excel打开后显示的表
print(type(sheet_active))


print(excel.sheetnames)  # 获得所有表名组成的列表

sheet1 = excel[excel.sheetnames[0]]  # 通过传入表名称获得工作表，得到一个工作表对象
print(type(sheet1))

print(sheet1.title)  # 得到工作表对象的表名


'''以下单元格均指有数据的单元格'''
# 获取sheet1单元格A1的两种方法
A1 = sheet1['A1']
print(A1)
A1 = sheet1.cell(row=1, column=1)  # row为行，column为列
print(A1)
print(type(A1))  # 得到Cell对象
# 获取单元格内容, Cell对象的value属性
print(A1.value)
# A1.value = '唐 青 san'
print(A1.value)  # 通过给cell对象的value赋值可以改变单元格内容

sheet_tuple1 = sheet1['A1':'B2']  # 得到A1到B2矩形范围内所有单元格的元组，元素个数为行数，其中同行的单元格组成一个元组
print(sheet_tuple1)
sheet_tuple2 = sheet1['A']  # 得到A列所有单元格元组,
# print(sheet_tuple2)

sheet_tuple3 = sheet1[1]  # 得到第一行所有的单元格元组
print(sheet_tuple3)
sheet_tuple4 = sheet1[1:3]  # 得到1~3行所有单元格元组, 同行组成元组作为元素
print(sheet_tuple4)
# print(tuple(sheet1.rows))  # 得到所有单元格组成的元组，其中每行单元格形成元组作为元素
# print(tuple(sheet1.columns))  # 得到所有单元格组成的元组，其中每列单元格形成元组作为元素
print(type(sheet1.rows), type(sheet1.columns))  # ws2.rows和ws2.columns是generator对象，因此需要使用tuple进行“解析”下。
print(sheet1['A1'].row, sheet1['A3'].column, sheet1['A1'].coordinate)  # cell对象的row,column,coordinate可以获得cell的位置信息

print(openpyxl.utils.get_column_letter(1))  # 传入数字返回数字对应列的字母，如传入1返回A
print(openpyxl.utils.column_index_from_string('AA'))  # 与上个函数相反
print(sheet1.max_row)  # 获取工作表的行数
print(sheet1.max_column)  # 获取工作表的列数
# sheet1.cell(1, 1).hyperlink = 'https://www.bilibili.com/'  # 对单元格设置超链接

# cell的其他属性
print(sheet1['B1'].column_letter)   # 返回cell的字母列名
print(sheet1['B1'].coordinate)  # 返回cell的坐标
print(sheet1['B1'].col_idx)  # 返回cell所属列对应的数字
print(sheet1['B1'].encoding)  # 返回cell内容的编码方式
print(sheet1['B1'].data_type)  # 返回cell的数据类型（Excel支持2种数据类型：Numbers和Strings）
'''使用openpyxl进行工作的时候，当一个工作结束的时候我们需要进行Excel文件的保存操作：wb.save('Mytest.xlsx')。
这个保存唯一需要注意的是：文件是默认替换的。也就是说我们在保存文件的时候，openpyxl将进行替换而不发出告警'''
excel.save('第三次作业批改.xlsx')
excel.close()

'''创建一个表格'''
wb = openpyxl.Workbook()  # 创建一个Work对象

ws = wb.create_sheet('demo_sheet', 0)  # 在索引为0的位置创建一个名为demo_sheet的sheet页

ws.sheet_properties.tabColor = 'ff72BA'  # 对sheet页设置一个颜色（16位的RGB颜色）

# wb.remove(wb.active)  # 传入表名，删除指定工作表


'''字体样式'''
# 创建一个字体对象
demo_font = openpyxl.styles.Font(name='华文行楷',  # 字体
                                 size=11,  # 字号
                                 color='FF000000',  # 颜色（16进制）
                                 bold=True,  # 是否为粗体
                                 italic=True,  # 是否为斜体
                                 vertAlign='baseline',  # 垂直对齐方式
                                 underline='double',  # 下滑线
                                 strike=False  # 是否加删除线
                                 )
print(type(demo_font))
ws['A1'].value = '宋遗平'
ws['A1'].font = demo_font  # 设置单元格的字体样式

ws.merge_cells('A1:B2')  # 合并A1到B2单元格
ws.unmerge_cells('A1:B2')  # 拆分

wb.save('demo_excel.xlsx')

wb.close()
