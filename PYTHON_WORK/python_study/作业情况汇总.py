# coding=gbk
import openpyxl
from openpyxl import utils


# def work(name_list, *excel_name):
#     wb = openpyxl.Workbook()
#     sheet_name = '机电化工'
#     ws = wb.create_sheet(f'{sheet_name}', 0)
#     for i in range(len(name_list)):
#         ws[f'A{i + 1}'].value = name_list[i]
#     for excel in excel_name:
#         workbook = openpyxl.load_workbook(excel)
#         for a in tuple(workbook.active.rows):
#             a_name = a[0].value
#             a_grade = a[1].value
#             if a_name in name_list:
#                 ws[
#                     f'{openpyxl.utils.get_column_letter(excel_name.index(excel) + 2)}{name_list.index(a_name) + 1}'
#                 ] = a_grade
#             else:
#                 print(a_name, '出现错误')
#                 names = []
#                 for i in name_list_2128:
#                     if list(a_name)[0] == list(i)[0]:
#                         names.append(i)
#                 print(names)
#                 try:
#                     num = eval(input('请选择可疑值序号：'))
#                     suspicion_name = names[num - 1]
#                     ws[
#                         f'{openpyxl.utils.get_column_letter(excel_name.index(excel) + 2)}'
#                         f'{name_list.index(suspicion_name) + 1}'
#                     ] = a_grade
#                     print('添加完毕')
#                 except:
#                     print('已跳过')
#                     pass
#
#         print('------------------------------------')
#     wb.save('2128高等数学作业情况.xlsx')
#     wb.close()
#
#
# work(name_list_2128, '第二次作业收改 (1).xlsx', '第三次作业批改.xlsx', '第四次作业收改.xlsx')
